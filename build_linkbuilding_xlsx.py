# -*- coding: utf-8 -*-
"""בונה קובץ Excel של הזדמנויות בניית קישורים ל-kingchairs.co.il"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.properties import WorksheetProperties

wb = openpyxl.Workbook()

# ---------- עיצוב כללי ----------
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
CAT_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
CELL_FONT = Font(name="Arial", size=10)
LINK_FONT = Font(name="Arial", size=10, color="0563C1", underline="single")
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(horizontal="right", vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# צבע לפי רמת קושי
DIFF_FILL = {
    "קל":     PatternFill("solid", fgColor="C6EFCE"),
    "בינוני": PatternFill("solid", fgColor="FFEB9C"),
    "גבוה":   PatternFill("solid", fgColor="FFC7CE"),
}
# צבע לפי פוטנציאל
POT_FILL = {
    "גבוה מאוד": PatternFill("solid", fgColor="A9D08E"),
    "גבוה":      PatternFill("solid", fgColor="C6EFCE"),
    "בינוני":    PatternFill("solid", fgColor="FFEB9C"),
    "נמוך":      PatternFill("solid", fgColor="F2F2F2"),
}
# צבע כותרת קטגוריה
CAT_FILL = {
    1: PatternFill("solid", fgColor="2E75B6"),
    2: PatternFill("solid", fgColor="548235"),
    3: PatternFill("solid", fgColor="7030A0"),
    4: PatternFill("solid", fgColor="BF8F00"),
    5: PatternFill("solid", fgColor="C55A11"),
    6: PatternFill("solid", fgColor="843C0C"),
}

HEADERS = ["#", "שם האתר / הזדמנות", "כתובת URL", "סוג לינק / פעולה",
           "רמת קושי", "פוטנציאל SEO", "הערות ואסטרטגיה מומלצת"]

# ---------- הנתונים ----------
# כל שורה: (שם, url, סוג לינק, קושי, פוטנציאל, הערות)
categories = [
    ("1. בלוגים ואתרי עיצוב פנים ישראליים (תוכן אורגני / אזכור מותג)", 1, [
        ("WALLS Magazine – מגזין עיצוב פנים", "https://wallsmag.co.il/", "מאמר אורח / כתבת תדמית",
         "בינוני", "גבוה מאוד",
         "מגזין עיצוב מוביל עם DA גבוה. להציע כתבה על מגמות בריהוט מרופד יוקרתי / כיסאות ממוכנים, עם קרדיט וקישור בגוף הטקסט."),
        ("בניין ודיור (bvd.co.il)", "https://www.bvd.co.il/", "כתבת פרויקט / אזכור",
         "גבוה", "גבוה מאוד",
         "מגזין דגל לאדריכלות ועיצוב, סמכותיות גבוהה. לשלוח פרויקט לקוח מרשים המרוהט ע\"י King Chairs לפיצ'ר עריכתי."),
        ("מגזין פורטפוליו (prtfl.co.il)", "https://www.prtfl.co.il/", "כתבה עריכתית / השקת מוצר",
         "גבוה", "גבוה",
         "מגזין עיצוב, אמנות ואופנה יומי מ-2005. מתאים להשקת קולקציה / כיסא בהתאמה אישית כסיפור עיצובי."),
        ("YUNU DESIGN – בלוג עיצוב פנים", "https://yunu.design/", "מאמר אורח / שיתוף מומחה",
         "קל", "בינוני",
         "בלוג מעצבת פעיל, קל ליצירת קשר. להציע תוכן מומחה על בחירת ריפוד / ארגונומיה של כיסאות."),
        ("הבית של עידה (ida-home.co.il)", "https://www.ida-home.co.il/", "מאמר אורח / קולבורציה",
         "קל", "בינוני",
         "בלוג מעצבת פנים אישי. פוטנציאל לשת\"פ תוכן וקישור טבעי מתוך המלצת מוצר."),
        ("Shimrit Design – בלוג", "https://www.shimrit-design.co.il/", "אזכור בכתבה / שת\"פ",
         "קל", "בינוני",
         "כתבות על עיצוב דירות יוקרה 2025. להשתלב בכתבות ריהוט יוקרה כספק מומלץ."),
        ("Miss Garot – עיצוב והום סטיילינג", "https://www.missgarot.co.il/", "מאמר אורח",
         "קל", "בינוני",
         "תוכן על עיצוב מודע וקיימות. זווית: ריהוט מרופד איכותי כהשקעה לטווח ארוך."),
        ("ig-interiors – מגזין עיצוב הבית", "https://ig-interiors.co.il/", "מאמר אורח / כתבת ספק",
         "בינוני", "גבוה",
         "מגזין המרכז מאמרים ממעצבים מובילים. פלטפורמה טובה למאמר מקצועי עם קישור."),
    ]),
    ("2. אתרי ביקורות, דירוגים והשוואת מחירים", 2, [
        ("זאפ השוואת מחירים (ZAP)", "https://www.zap.co.il/", "רישום חנות / דף מותג",
         "בינוני", "גבוה מאוד",
         "אתר השוואת המחירים הגדול בישראל. לפתוח חנות מאומתת בקטגוריית ריהוט לבית + לצבור חוות דעת. תנועה מסחרית גבוהה."),
        ("זאפ – ריהוט לבית (קטגוריה)", "https://www.zap.co.il/world.aspx?world=home-furnitures",
         "רישום מוצרים בקטגוריה", "בינוני", "גבוה",
         "להזין את קטלוג הכיסאות/הספות עם קישור לדף המוצר באתר. חשיפה לקונים בכוונת רכישה."),
        ("מדרג – דירוג בעלי מקצוע", "https://www.midrag.co.il/", "פרופיל עסק + ביקורות",
         "קל", "גבוה",
         "לפתוח כרטיס עסק, לעודד לקוחות להשאיר ביקורות. קישור לאתר + סיגנל אמון (E-E-A-T)."),
        ("יד2 (Yad2) – עסקים/יד שנייה", "https://www.yad2.co.il/", "דף עסק / פרסום מוצרים",
         "קל", "בינוני",
         "פרסום ריהוט + פתיחת עמוד עסק. תנועה עצומה, קישור nofollow אך ערך חשיפה גבוה."),
        ("Toplink – אינדקס עסקים וביקורות", "https://toplink.co.il/", "רישום עסק",
         "קל", "נמוך",
         "אינדקס עסקים קל לרישום. קישור בסיסי לבניית פרופיל ציטוטים (NAP)."),
    ]),
    ("3. פורומים וקבוצות פייסבוק לעיצוב הבית", 3, [
        ("פשפשוק – ריהוט ועיצוב הבית", "https://www.facebook.com/groups/Pishpeshuk.Yesh.Inyan/",
         "שיתוף תוכן / מיתוג", "קל", "בינוני",
         "רשת קהילות ענק (מאות אלפי חברים). לשתף תוכן ערך (לא ספאם) ולבנות מודעות מותג + הפניות."),
        ("עיצוב הבית – רעיונות, טיפים ויעוץ", "https://www.facebook.com/groups/1318902364786686/",
         "מענה מומחה / הפניה", "קל", "בינוני",
         "קבוצת ייעוץ פעילה. לענות כמומחה ריהוט ולהפנות לתוכן מקצועי באתר."),
        ("משוגעים על עיצוב נורדי", "https://www.facebook.com/groups/590692554472250/",
         "שיתוף מוצר / השראה", "קל", "בינוני",
         "31K+ חברים אוהבי עיצוב סקנדינבי. מתאים להצגת כיסאות בעיצוב נקי/מודרני."),
        ("עיצוב פנים וטיפים לעיצוב הבית", "https://www.facebook.com/groups/1391293584464335",
         "מענה מומחה / תוכן", "קל", "בינוני",
         "קהילת עיצוב גדולה. בניית סמכות ותנועת רפרל לאתר."),
        ("עיצוב הבית , רעיונות , טיפים ויעוץ (Zehava Avitan)", "https://www.facebook.com/groups/263616147051632/",
         "שיתוף השראה / הפניה", "קל", "בינוני",
         "51K+ חברים בניהול מעצבת. פלטפורמה להצגת פרויקטים מרוהטים."),
    ]),
    ("4. קטלוגים, אינדקסים ועסקים מקומיים (Citations / Local SEO)", 4, [
        ("דפי זהב (d.co.il)", "https://www.d.co.il/", "כרטיס עסק + קטגוריית רהיטים",
         "קל", "גבוה מאוד",
         "אינדקס העסקים המוביל. רישום בקטגוריית 'חנויות רהיטים' + עיר. קריטי ל-Local SEO וציטוט NAP עקבי."),
        ("דפי זהב – חנויות רהיטים לפי עיר", "https://www.d.co.il/h-c45870-e0-p0-l0-city5000/recommended/",
         "רישום בקטגוריה + ביקורות", "קל", "גבוה",
         "להירשם בקטגוריית הרהיטים בעיר הפעילות ולצבור המלצות למיקום גבוה."),
        ("וואלה עסקים / אינדקס", "https://www.walla.co.il/", "רישום עסק / כתבת יח\"צ",
         "בינוני", "גבוה",
         "פורטל ותיק עם DA גבוה מאוד. רישום עסק + אפשרות כתבת יח\"צ ממומנת עם קישור."),
        ("Google Business Profile (עסק שלי)", "https://www.google.com/business/", "פרופיל עסק מקומי",
         "קל", "גבוה מאוד",
         "בסיס ה-Local SEO. פרופיל מלא, תמונות, אתר וביקורות — משפיע ישירות על מפות וחיפוש מקומי."),
        ("prosites – פורטל עסקים", "https://prosites.co.il/", "רישום עסק חינם",
         "קל", "נמוך",
         "אינדקס עסקים לפרסום חינם. לינק בסיסי לחיזוק פרופיל ציטוטים."),
        ("EasyIndex / אינדקסים ישראליים נוספים", "https://www.easy.co.il/", "רישום עסק",
         "קל", "נמוך",
         "אינדקס עסקים כללי. חלק מאסטרטגיית ציטוטים רחבה (מגוון מקורות)."),
    ]),
    ("5. הזדמנויות Guest Post באתרי אדריכלות ועיצוב", 5, [
        ("Xnet / ynet – עיצוב ואדריכלות", "https://www.ynet.co.il/architecture", "כתבה עריכתית / יח\"צ",
         "גבוה", "גבוה מאוד",
         "אתר החדשות המוביל, סמכותיות עצומה. לפנות דרך יח\"צ/מערכת עם סיפור עיצובי או פרויקט חריג לפיצ'ר."),
        ("קינן אדריכלים – בלוג", "https://keinan-arch.com/category/%D7%91%D7%9C%D7%95%D7%92/",
         "פוסט אורח / שת\"פ תוכן", "בינוני", "גבוה",
         "בלוג משרד אדריכלים מקצועי. להציע תוכן משותף על אינטגרציית ריהוט בפרויקטים."),
        ("דניאל וקנין – בלוג עיצוב ואדריכלות", "https://www.daniel-vaknin.co.il/", "מאמר אורח",
         "בינוני", "בינוני",
         "בלוג על עיצוב פנטהאוזים וטכנולוגיה. זווית מצוינת: ריהוט ממוכן/סמארט לבית יוקרה."),
        ("שרון דוד – בלוג עיצוב", "https://sharondavid.co.il/", "מאמר אורח / קולבורציה",
         "קל", "בינוני",
         "בלוג עיצוב וינטג'/יוקרה. פוטנציאל לפוסט משותף עם קישור טבעי."),
        ("ronit-tal / doritsela / m2arc – משרדי אדריכלות", "https://www.ronit-tal.co.il/",
         "שת\"פ פרויקט / אזכור הדדי", "בינוני", "בינוני",
         "משרדי אדריכלות עם בלוגים. הצעת ריהוט לפרויקטים בתמורה לקרדיט וקישור הדדי."),
    ]),
    ("6. שיתופי פעולה עם מעצבי פנים (Partnerships / Referral)", 6, [
        ("רשת מעצבי פנים עצמאיים (בלוגים אישיים)", "—", "עמוד ספקים מומלצים / קישור הדדי",
         "בינוני", "גבוה",
         "לבנות תוכנית 'מעצב מועדף': מעצבים מפנים לקוחות ומקשרים לאתר מעמוד 'ספקים/שותפים'. קישורי dofollow איכותיים והקשריים."),
        ("2M עיצוב פנים ואדריכלות (m2arc)", "https://m2arc.com/", "אזכור בפרויקט / קישור הדדי",
         "בינוני", "גבוה",
         "משרד עם עמוד עיטורים/פרויקטים. לספק ריהוט לפרויקט ולקבל קרדיט + קישור בעמוד הפרויקט."),
        ("איגוד / קהילות מעצבי פנים בישראל", "—", "חסות / תוכן מקצועי / דירקטורי",
         "בינוני", "בינוני",
         "חברות בקהילות מקצועיות ודירקטורי ספקים. חשיפה למקבלי החלטות + קישורים מדפי חברים."),
        ("סטודיו הום סטיילינג (Miss Garot וכד')", "https://www.missgarot.co.il/", "קולבורציית תוכן / staging",
         "קל", "בינוני",
         "שת\"פ Home-Staging: הצבת כיסאות King Chairs בפרויקטי צילום תמורת קרדיט וקישור."),
        ("משפיעני עיצוב באינסטגרם/יוטיוב (עם אתר)", "—", "שת\"פ ממומן / ביקורת מוצר",
         "בינוני", "גבוה",
         "מיקרו-משפיענים בתחום עיצוב הבית שיש להם גם אתר/בלוג. ביקורת מוצר + קישור מהבלוג (לא רק סושיאל)."),
    ]),
]

ws = wb.active
ws.title = "הזדמנויות לינקים"
ws.sheet_view.rightToLeft = True

# רוחב עמודות
widths = [5, 34, 46, 24, 11, 13, 60]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

# ---------- כותרת ראשית ----------
ws.merge_cells("A1:G1")
c = ws["A1"]
c.value = "King Chairs (kingchairs.co.il) – מפת הזדמנויות בניית קישורים אורגניים (Link Building) | ריהוט מרופד יוקרתי, ריהוט ממוכן וכיסאות בהתאמה אישית"
c.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
c.fill = PatternFill("solid", fgColor="203864")
c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[1].height = 46

ws.merge_cells("A2:G2")
c = ws["A2"]
c.value = "נכון ל-יולי 2026 | דירוג קושי: קל=רישום עצמי/פנייה פשוטה · בינוני=דורש פיץ' ותוכן · גבוה=דורש קשר עורכי/יח\"צ.  פוטנציאל SEO = השפעה משוקללת (סמכות דומיין + רלוונטיות + תנועה)."
c.font = Font(name="Arial", size=9, italic=True, color="595959")
c.fill = PatternFill("solid", fgColor="D9E1F2")
c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[2].height = 30

row = 4
row_counter = 0
for cat_title, cat_id, items in categories:
    # כותרת קטגוריה
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    c = ws.cell(row=row, column=1, value=cat_title)
    c.font = CAT_FONT
    c.fill = CAT_FILL[cat_id]
    c.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[row].height = 24
    row += 1

    # שורת כותרות עמודות
    for col, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = 22
    row += 1

    # שורות נתונים
    for name, url, link_type, diff, pot, notes in items:
        row_counter += 1
        vals = [row_counter, name, url, link_type, diff, pot, notes]
        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=row, column=col, value=v)
            c.border = BORDER
            if col == 1:
                c.alignment = CENTER
                c.font = CELL_FONT
            elif col == 3 and url != "—":
                c.hyperlink = url
                c.font = LINK_FONT
                c.alignment = WRAP
            elif col == 5:
                c.alignment = CENTER
                c.font = Font(name="Arial", size=10, bold=True)
                c.fill = DIFF_FILL.get(diff, PatternFill())
            elif col == 6:
                c.alignment = CENTER
                c.font = Font(name="Arial", size=10, bold=True)
                c.fill = POT_FILL.get(pot, PatternFill())
            else:
                c.alignment = WRAP
                c.font = CELL_FONT
        row += 1
    row += 1  # רווח בין קטגוריות

# ---------- גיליון מקרא ----------
ws2 = wb.create_sheet("מקרא ומתודולוגיה")
ws2.sheet_view.rightToLeft = True
ws2.column_dimensions["A"].width = 22
ws2.column_dimensions["B"].width = 90
legend = [
    ("שדה", "הסבר"),
    ("רמת קושי – קל", "ניתן לבצע עצמאית: רישום באינדקס, פתיחת כרטיס עסק, שיתוף בקבוצה, פנייה ישירה לבלוגר קטן."),
    ("רמת קושי – בינוני", "דורש הכנת תוכן איכותי (מאמר אורח), פיץ' מסודר או משא ומתן על שת\"פ."),
    ("רמת קושי – גבוה", "דורש קשר עורכי/יחסי ציבור, לרוב אתרים סמכותיים (ynet, בניין ודיור) — עשוי לדרוש תקציב/PR."),
    ("פוטנציאל – גבוה מאוד", "השפעת SEO מקסימלית: דומיין סמכותי מאוד + רלוונטיות גבוהה + תנועה מסחרית."),
    ("פוטנציאל – גבוה", "השפעה חזקה על דירוג/תנועה, רלוונטי מאוד לנישת ריהוט יוקרה."),
    ("פוטנציאל – בינוני", "תורם לפרופיל הקישורים והמודעות, השפעה ישירה מתונה."),
    ("פוטנציאל – נמוך", "ערך בעיקר לבניית ציטוטים (NAP) ועקביות, לא לינק \"חזק\"."),
    ("", ""),
    ("המלצות עבודה", ""),
    ("1", "לתעדף: קודם ה-Local SEO (דפי זהב, Google Business, מדרג, זאפ) — בסיס אמון וציטוטים מהיר וזול."),
    ("2", "במקביל להתחיל פיצ'ים למגזינים (WALLS, בניין ודיור, פורטפוליו, Xnet) — לוקח זמן אך בעל ערך גבוה מאוד."),
    ("3", "לבנות תוכנית 'מעצב מועדף' לקישורים הדדיים איכותיים ומתמשכים ממעצבי פנים."),
    ("4", "לגוון עוגני קישור (anchor text): שם מותג, מונחי נישה (\"כיסא ממוכן\", \"ריהוט מרופד בהתאמה אישית\"), וכתובת URL."),
    ("5", "להימנע מספאם בקבוצות פייסבוק — לתת ערך אמיתי; קישורי סושיאל הם nofollow אך תורמים לתנועה ומודעות."),
    ("6", "לעקוב אחר קישורים שהושגו בגיליון נפרד (תאריך, סטטוס, dofollow/nofollow, עוגן) לניהול הקמפיין."),
]
for r, (a, b) in enumerate(legend, start=1):
    ca = ws2.cell(row=r, column=1, value=a)
    cb = ws2.cell(row=r, column=2, value=b)
    if r == 1 or a == "המלצות עבודה":
        ca.font = Font(name="Arial", bold=True, color="FFFFFF")
        cb.font = Font(name="Arial", bold=True, color="FFFFFF")
        ca.fill = HEADER_FILL
        cb.fill = HEADER_FILL
    else:
        ca.font = Font(name="Arial", bold=True, size=10)
        cb.font = CELL_FONT
    ca.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
    cb.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)

# הקפאת שורות עליונות בגיליון הראשי
ws.freeze_panes = "A4"

out = "/home/user/kas-hamelech-meta/KingChairs_LinkBuilding_Opportunities.xlsx"
wb.save(out)
print("saved:", out)
print("total opportunities:", row_counter)
